from rest_framework import generics, status, viewsets, filters
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from documents.models import (
    ISOClause, Tag, Document, ApprovalWorkflow, 
    ChangeRequest, DocumentTemplate, Record, DocumentFolder
)
from documents.serializers import (
    ISOClauseSerializer, TagSerializer, DocumentSerializer, 
    ApprovalWorkflowSerializer, ChangeRequestSerializer, 
    DocumentTemplateSerializer, CreateDocumentFromTemplateSerializer,
    RecordSerializer, RecordApprovalSerializer, DocumentFolderSerializer
)
from quickreports.models import QuickReport
from quickreports.serializers import QuickReportSerializer, QuickReportReviewSerializer
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.exceptions import PermissionDenied
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.shortcuts import get_object_or_404
from django.db.models import Q, Sum, Count, Avg
from collections import defaultdict
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from rest_framework.decorators import action
from .permissions import IsHSSEManager, LegalCompliancePermission, PPEManagementPermission, AuditManagementPermission, RiskManagementPermission
from legals.models import (
    LawCategory, LawResource, LawResourceChange,
    LegalRegisterEntry, LegalRegisterComment, LegalRegisterDocument, Position, LegislationTracker
)
from legals.serializers import (
    LawCategorySerializer, LawResourceSerializer, LawResourceChangeSerializer,
    LegalRegisterEntrySerializer, LegalRegisterCommentSerializer, LegalRegisterDocumentSerializer,
    PositionSerializer, LegislationTrackerSerializer
)
from ppes.models import (
    PPECategory, Vendor, PPEPurchase, PPEInventory, PPEIssue, 
    PPERequest, PPEDamageReport, PPETransfer, PPEReturn, PPEPurchaseReceipt
)
from ppes.serializers import (
    PPECategorySerializer, VendorSerializer, PPEPurchaseSerializer, PPEInventorySerializer,
    PPEIssueSerializer, PPERequestSerializer, PPERequestApprovalSerializer,
    PPEDamageReportSerializer, PPEDamageReportReviewSerializer,
    PPETransferSerializer, PPETransferApprovalSerializer, PPEReturnSerializer,
    PPEStockPositionSerializer, PPEMovementSerializer, PPEMostRequestedSerializer,
    PPECostAnalysisSerializer, PPEExpiryAlertSerializer, PPELowStockAlertSerializer,
    PPEUserStockSerializer, BulkPPEIssueSerializer, BulkPPERequestApprovalSerializer,
    PPEPurchaseReceiptSerializer
)
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.db import connection
from django.core.cache import cache
from django.conf import settings
import os

User = get_user_model()

# =============================
# Generic API Views
# =============================
class DocumentListCreateAPIView(generics.ListCreateAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Document.objects.all()
        
        # Handle search filtering
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(content__icontains=search)
            )
        
        # Handle category filtering (map folder value to document_type)
        category = self.request.query_params.get('category', None)
        if category:
            # Map SYSTEM_DOCUMENT (folder value) to SYSTEM DOCUMENT (model value)
            document_type = 'SYSTEM DOCUMENT' if category == 'SYSTEM_DOCUMENT' else category
            queryset = queryset.filter(document_type=document_type)
            
            # For FORM folder: Only show templates (not records)
            # Templates are documents with document_type='FORM' that are NOT records
            # Records are stored separately in the Record model
            if document_type == 'FORM':
                # Ensure we only show actual form templates (documents), not records
                # Records are in a separate model, so this filter ensures we only get Document instances
                # which are templates by definition when document_type='FORM'
                queryset = queryset.filter(is_template=False)  # Documents are not templates (templates are in DocumentTemplate model)
        
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class DocumentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Document.objects.all()
    serializer_class = DocumentSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'  # This will use the UUID field

    def get_object(self):
        obj = get_object_or_404(self.get_queryset(), id=self.kwargs["pk"])
        self.check_object_permissions(self.request, obj)
        return obj

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


# =============================
# Document Folder API Views
# =============================

class DocumentFolderListCreateAPIView(generics.ListCreateAPIView):
    """
    List all folders or create a new folder.
    Only HSSE Manager and Superadmins can create folders.
    """
    queryset = DocumentFolder.objects.filter(is_active=True)
    serializer_class = DocumentFolderSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """Return all active folders, ordered by name."""
        return DocumentFolder.objects.filter(is_active=True).order_by('name')
    
    def get_permissions(self):
        """Only HSSE Manager/Superadmin can create folders."""
        if self.request.method == 'POST':
            user = self.request.user
            if not (user.position == 'HSSE MANAGER' or user.is_superuser):
                return [IsHSSEManager()]  # This will raise PermissionDenied
        return super().get_permissions()
    
    def perform_create(self, serializer):
        """Set created_by and validate permissions."""
        user = self.request.user
        if not (user.position == 'HSSE MANAGER' or user.is_superuser):
            raise PermissionDenied("Only HSSE Managers and Superadmins can create folders.")
        serializer.save(created_by=user)
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class DocumentFolderRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """
    Retrieve, update, or delete a folder.
    Only HSSE Manager and Superadmins can update/delete.
    Folders can only be deleted if they are empty.
    """
    queryset = DocumentFolder.objects.all()
    serializer_class = DocumentFolderSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    
    def get_object(self):
        obj = get_object_or_404(self.get_queryset(), id=self.kwargs["pk"])
        self.check_object_permissions(self.request, obj)
        return obj
    
    def get_permissions(self):
        """Only HSSE Manager/Superadmin can update/delete folders."""
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            user = self.request.user
            if not (user.position == 'HSSE MANAGER' or user.is_superuser):
                return [IsHSSEManager()]  # This will raise PermissionDenied
        return super().get_permissions()
    
    def perform_update(self, serializer):
        """Validate permissions before update."""
        user = self.request.user
        if not (user.position == 'HSSE MANAGER' or user.is_superuser):
            raise PermissionDenied("Only HSSE Managers and Superadmins can update folders.")
        serializer.save()
    
    def perform_destroy(self, instance):
        """Validate permissions and ensure folder is empty before deletion."""
        user = self.request.user
        if not (user.position == 'HSSE MANAGER' or user.is_superuser):
            raise PermissionDenied("Only HSSE Managers and Superadmins can delete folders.")
        
        # Check if folder is empty
        if not instance.is_empty():
            document_count = instance.get_document_count()
            raise PermissionDenied(
                f"Cannot delete folder '{instance.name}' because it contains {document_count} document(s). "
                "Please delete all documents in this folder first."
            )
        
        # Soft delete by setting is_active=False
        instance.is_active = False
        instance.save()
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class ISOClauseListCreateAPIView(generics.ListCreateAPIView):
    queryset = ISOClause.objects.all()
    serializer_class = ISOClauseSerializer


class TagListCreateAPIView(generics.ListCreateAPIView):
    queryset = Tag.objects.all()
    serializer_class = TagSerializer


class ChangeRequestListCreateAPIView(generics.ListCreateAPIView):
    queryset = ChangeRequest.objects.all()
    serializer_class = ChangeRequestSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)


class ChangeRequestRetrieveUpdateAPIView(generics.RetrieveUpdateAPIView):
    queryset = ChangeRequest.objects.all()
    serializer_class = ChangeRequestSerializer


class ApprovalWorkflowListCreateAPIView(generics.ListCreateAPIView):
    queryset = ApprovalWorkflow.objects.all()
    serializer_class = ApprovalWorkflowSerializer
    

# =============================
# Approval Action Views
# =============================
class SubmitDocumentAPIView(APIView):
    def post(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            if not doc.is_editable(request.user):
                return Response({"error": "You don't have permission to submit this document."}, status=403)
            
            doc.transition_to('VERIFICATION', request.user)
            return Response({'status': 'Document submitted for verification'})
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class VerifyDocumentAPIView(APIView):
    def post(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            if not doc.is_verifiable(request.user):
                return Response({"error": "You don't have permission to verify this document."}, status=403)
            
            doc.transition_to('APPROVAL', request.user)
            doc.verified_by = request.user
            doc.verified_at = timezone.now()
            doc.save()
            return Response({'status': 'Document verified, pending approval'})
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class ApproveDocumentAPIView(APIView):
    def post(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            if not doc.is_approvable(request.user):
                return Response({"error": "You don't have permission to approve this document."}, status=403)
            
            doc.transition_to('APPROVED', request.user)
            doc.approved_by = request.user
            doc.approved_at = timezone.now()
            doc.revision_number += 1
            doc.save()
            return Response({'status': 'Document approved'})
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class RejectDocumentAPIView(APIView):
    def post(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            if not doc.is_rejectable(request.user):
                return Response({"error": "You don't have permission to reject this document."}, status=403)
            
            reason = request.data.get('reason', '')
            if not reason:
                return Response({"error": "Rejection reason is required."}, status=400)
            
            doc.transition_to('REJECTED', request.user, comment=reason)
            doc.rejection_reason = reason
            doc.save()
            return Response({'status': 'Document rejected'})
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class DocumentWorkflowHistoryAPIView(APIView):
    def get(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            workflow_history = doc.get_workflow_history()
            serializer = ApprovalWorkflowSerializer(workflow_history, many=True)
            return Response(serializer.data)
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)


class CreateNewVersionAPIView(APIView):
    def post(self, request, pk):
        try:
            doc = Document.objects.get(pk=pk)
            if not doc.is_editable(request.user):
                return Response({"error": "You don't have permission to create a new version."}, status=403)
            
            doc.create_new_version()
            return Response({
                'status': 'New version created',
                'version': doc.get_current_version()
            })
        except Document.DoesNotExist:
            return Response({"error": "Document not found."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


# =============================
# Template Management Views
# =============================
class DocumentTemplateListCreateAPIView(generics.ListCreateAPIView):
    queryset = DocumentTemplate.objects.filter(is_active=True)
    serializer_class = DocumentTemplateSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class DocumentTemplateRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = DocumentTemplate.objects.all()
    serializer_class = DocumentTemplateSerializer
    
    def perform_update(self, serializer):
        # Only allow updates if user is the creator or has appropriate permissions
        if self.request.user != serializer.instance.created_by and not self.request.user.is_staff:
            raise PermissionDenied("You don't have permission to update this template")
        serializer.save()


class CreateDocumentFromTemplateAPIView(APIView):
    def post(self, request, template_id):
        try:
            template = DocumentTemplate.objects.get(id=template_id, is_active=True)
            serializer = CreateDocumentFromTemplateSerializer(
                data=request.data,
                context={'template': template}
            )
            serializer.is_valid(raise_exception=True)
            
            # Create document from template
            document = template.create_document_from_template(
                user=request.user,
                title=serializer.validated_data['title'],
                description=serializer.validated_data.get('description', '')
            )
            
            # Update document metadata if provided
            if 'metadata' in serializer.validated_data:
                document.metadata.update(serializer.validated_data['metadata'])
                document.save()
            
            return Response({
                'status': 'Document created successfully',
                'document_id': document.id
            })
            
        except DocumentTemplate.DoesNotExist:
            return Response({"error": "Template not found or inactive"}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class ApproveTemplateAPIView(APIView):
    def post(self, request, template_id):
        try:
            template = DocumentTemplate.objects.get(id=template_id)
            if not request.user.position == 'MD':
                return Response({"error": "Only MD can approve templates"}, status=403)
            
            template.approved_by = request.user
            template.approved_at = timezone.now()
            template.save()
            
            return Response({'status': 'Template approved successfully'})
            
        except DocumentTemplate.DoesNotExist:
            return Response({"error": "Template not found"}, status=404)


# =============================
# Document Dashboard API View
# =============================
class DocumentDashboardAPIView(APIView):
    """API endpoint for document management dashboard metrics."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Get basic document counts
            total_documents = Document.objects.count()
            
            # Get documents by status
            status_counts = Document.objects.values('status').annotate(count=Count('id'))
            status_dict = {item['status']: item['count'] for item in status_counts}
            
            # Get documents by type
            type_counts = Document.objects.values('document_type').annotate(count=Count('id'))
            type_dict = {item['document_type']: item['count'] for item in type_counts}
            
            # Get change requests count
            change_requests_count = ChangeRequest.objects.count()
            
            # Get recent activities (last 10 workflow entries)
            recent_activities = ApprovalWorkflow.objects.select_related(
                'document', 'performed_by'
            ).order_by('-timestamp')[:10]
            
            # Format recent activities
            activities = []
            for activity in recent_activities:
                activities.append({
                    'id': activity.id,
                    'document_title': activity.document.title,
                    'action': activity.action,
                    'performed_by': activity.performed_by.get_full_name() if activity.performed_by else 'System',
                    'created_at': activity.timestamp,
                    'comment': activity.comment
                })
            
            # Calculate percentages for document types
            document_types = {
                'policy': type_dict.get('POLICY', 0),
                'system_document': type_dict.get('SYSTEM DOCUMENT', 0),
                'procedure': type_dict.get('PROCEDURE', 0),
                'form': type_dict.get('FORM', 0),
                'ssow': type_dict.get('SSOW', 0),
                'other': type_dict.get('OTHER', 0)
            }
            
            # Calculate status percentages
            status_breakdown = {
                'draft': status_dict.get('DRAFT', 0),
                'hsse_review': status_dict.get('HSSE_REVIEW', 0),
                'ops_review': status_dict.get('OPS_REVIEW', 0),
                'md_approval': status_dict.get('MD_APPROVAL', 0),
                'approved': status_dict.get('APPROVED', 0),
                'rejected': status_dict.get('REJECTED', 0)
            }
            
            # Calculate pending approvals (documents in review stages)
            pending_approvals = (
                status_dict.get('HSSE_REVIEW', 0) + 
                status_dict.get('OPS_REVIEW', 0) + 
                status_dict.get('MD_APPROVAL', 0)
            )
            
            response_data = {
                'metrics': {
                    'total_documents': total_documents,
                    'pending_approvals': pending_approvals,
                    'change_requests': change_requests_count,
                    'approved_documents': status_dict.get('APPROVED', 0),
                    'rejected_documents': status_dict.get('REJECTED', 0),
                    'draft_documents': status_dict.get('DRAFT', 0),
                },
                'document_types': document_types,
                'status_breakdown': status_breakdown,
                'recent_activities': activities
            }
            
            return Response(response_data)
            
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch dashboard data: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================
# Document Review Schedule Dashboard API View
# =============================
class DocumentReviewScheduleAPIView(APIView):
    """API endpoint for document review schedule dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from datetime import date, timedelta
        from django.db.models import Q
        
        try:
            today = date.today()
            thirty_days = today + timedelta(days=30)
            sixty_days = today + timedelta(days=60)
            ninety_days = today + timedelta(days=90)
            
            # Get all approved documents with review dates
            documents = Document.objects.filter(
                status='APPROVED',
                is_obsolete=False
            ).select_related('approved_by', 'created_by').prefetch_related('iso_clauses')
            
            overdue_docs = []
            due_soon_docs = []  # Within 30 days
            upcoming_docs = []  # 31-60 days
            later_docs = []  # 61-90 days
            no_review_date = []
            
            for doc in documents:
                if not doc.next_review_date:
                    no_review_date.append({
                        'id': str(doc.id),
                        'title': doc.title,
                        'document_type': doc.document_type,
                        'version': doc.version,
                        'revision_number': float(doc.revision_number),
                        'approved_at': doc.approved_at.isoformat() if doc.approved_at else None,
                        'approved_by': doc.approved_by.get_full_name() if doc.approved_by else None,
                        'status': doc.status,
                        'document_classification': doc.document_classification,
                        'days_until_review': None,
                        'is_overdue': False,
                    })
                    continue
                
                days_until = (doc.next_review_date - today).days
                is_overdue = days_until < 0
                
                doc_data = {
                    'id': str(doc.id),
                    'title': doc.title,
                    'document_type': doc.document_type,
                    'version': doc.version,
                    'revision_number': float(doc.revision_number),
                    'next_review_date': doc.next_review_date.isoformat(),
                    'expiry_date': doc.expiry_date.isoformat() if doc.expiry_date else None,
                    'approved_at': doc.approved_at.isoformat() if doc.approved_at else None,
                    'approved_by': doc.approved_by.get_full_name() if doc.approved_by else None,
                    'status': doc.status,
                    'document_classification': doc.document_classification,
                    'days_until_review': days_until,
                    'is_overdue': is_overdue,
                }
                
                if is_overdue:
                    overdue_docs.append(doc_data)
                elif days_until <= 30:
                    due_soon_docs.append(doc_data)
                elif days_until <= 60:
                    upcoming_docs.append(doc_data)
                elif days_until <= 90:
                    later_docs.append(doc_data)
            
            # Sort by urgency (overdue first, then by days until review)
            overdue_docs.sort(key=lambda x: x['days_until_review'])
            due_soon_docs.sort(key=lambda x: x['days_until_review'])
            upcoming_docs.sort(key=lambda x: x['days_until_review'])
            later_docs.sort(key=lambda x: x['days_until_review'])
            
            # Statistics
            total_docs = documents.count()
            total_with_review_dates = total_docs - len(no_review_date)
            overdue_count = len(overdue_docs)
            due_soon_count = len(due_soon_docs)
            
            response_data = {
                'statistics': {
                    'total_documents': total_docs,
                    'with_review_dates': total_with_review_dates,
                    'without_review_dates': len(no_review_date),
                    'overdue': overdue_count,
                    'due_soon': due_soon_count,
                    'upcoming': len(upcoming_docs),
                    'later': len(later_docs),
                },
                'overdue': overdue_docs[:50],  # Limit to 50 for performance
                'due_soon': due_soon_docs[:50],
                'upcoming': upcoming_docs[:50],
                'later': later_docs[:50],
                'no_review_date': no_review_date[:50],
            }
            
            return Response(response_data)
            
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch review schedule: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ChangeRequestApproveAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        change_request = get_object_or_404(ChangeRequest, pk=pk)
        if not change_request.can_be_approved_by(request.user):
            return Response({'error': 'Only HSSE Manager can approve change requests.'}, status=403)
        response = request.data.get('response', '')
        try:
            change_request.approve(request.user, response)
            return Response({'status': 'Change request approved.'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class ChangeRequestRejectAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        change_request = get_object_or_404(ChangeRequest, pk=pk)
        if not change_request.can_be_rejected_by(request.user):
            return Response({'error': 'Only HSSE Manager can reject change requests.'}, status=403)
        response = request.data.get('response', '')
        try:
            change_request.reject(request.user, response)
            return Response({'status': 'Change request rejected.'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class SubmitForOpsReviewAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        doc = get_object_or_404(Document, id=pk)
        if not (doc.status == 'DRAFT' and request.user.position == 'HSSE MANAGER'):
            return Response({'error': 'Only HSSE Manager can submit for OPS review.'}, status=403)
        try:
            doc.transition_to('OPS_REVIEW', request.user)
            return Response({'status': 'Document submitted for OPS review.'})
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class OpsReviewAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, pk):
        doc = get_object_or_404(Document, id=pk)
        action = request.data.get('action')  # 'approve' or 'reject'
        comment = request.data.get('comment', '')
        if not doc.status == 'OPS_REVIEW':
            return Response({'error': 'Document is not in OPS review stage.'}, status=400)
        if request.user.position != 'OPS MANAGER':
            return Response({'error': 'Only OPS Manager can review at this stage.'}, status=403)
        try:
            if action == 'approve':
                doc.transition_to('MD_APPROVAL', request.user, comment)
                return Response({'status': 'Document approved and sent to MD for approval.'})
            elif action == 'reject':
                doc.transition_to('REJECTED', request.user, comment)
                return Response({'status': 'Document rejected by OPS Manager.'})
            else:
                return Response({'error': 'Invalid action.'}, status=400)
        except Exception as e:
            return Response({'error': str(e)}, status=400)


class MDApprovalAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        doc = get_object_or_404(Document, pk=pk)
        action_type = request.data.get('action') # 'approve' or 'reject'
        comment = request.data.get('comment', '')

        if not doc.is_md_approvable(request.user):
            raise PermissionDenied("You do not have permission to review this document.")

        if action_type == 'approve':
            doc.transition_to('APPROVED', request.user, comment)
            doc.approved_by = request.user
            doc.approved_at = timezone.now()
            doc.save()
            return Response({'status': 'Document approved by MD.'})
        elif action_type == 'reject':
            doc.transition_to('REJECTED', request.user, comment)
            doc.rejection_reason = comment
            doc.save()
            return Response({'status': 'Document rejected by MD.'})
        else:
            return Response({'error': 'Invalid action.'}, status=status.HTTP_400_BAD_REQUEST)


# =================================
# Quick Reports Management
# =================================

class QuickReportViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing Quick Reports (Accidents, Near Misses, Potential Incidents, Non-Conformities).
    
    Features:
    - All users can submit reports
    - HSSE Managers/Admins can approve/reject
    - On approval, automatically creates a Record
    - Email and in-app notifications
    """
    queryset = QuickReport.objects.all().order_by('-created_at')
    serializer_class = QuickReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['report_type', 'severity', 'status', 'reported_by']
    search_fields = ['report_number', 'title', 'description', 'location']
    ordering_fields = ['created_at', 'incident_date', 'severity']

    def get_queryset(self):
        user = self.request.user
        queryset = QuickReport.objects.all() if (user.position == 'HSSE MANAGER' or user.is_superuser) else QuickReport.objects.filter(reported_by=user)
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(reported_by=self.request.user)

    @action(detail=True, methods=['post'], permission_classes=[IsHSSEManager])
    def review(self, request, pk=None):
        """Review (approve/reject) a quick report."""
        quick_report = self.get_object()
        serializer = QuickReportReviewSerializer(data=request.data)
        
        if serializer.is_valid():
            action = serializer.validated_data['action']
            comments = serializer.validated_data.get('comments', '')
            rejection_reason = serializer.validated_data.get('rejection_reason', '')
            
            try:
                if action == 'approve':
                    quick_report.approve(request.user, comments)
                    return Response({
                        'status': 'Quick report approved',
                        'message': f'Report {quick_report.report_number} approved and filed as a record.',
                        'record_number': quick_report.created_record.record_number if quick_report.created_record else None
                    }, status=status.HTTP_200_OK)
                else:
                    quick_report.reject(request.user, rejection_reason)
                    return Response({
                        'status': 'Quick report rejected',
                        'message': f'Report {quick_report.report_number} rejected. Submitter has been notified.'
                    }, status=status.HTTP_200_OK)
            except ValueError as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def statistics(self, request):
        """Get quick report statistics."""
        user = request.user
        queryset = QuickReport.objects.all() if (user.position == 'HSSE MANAGER' or user.is_superuser) else QuickReport.objects.filter(reported_by=user)
        
        stats = {
            'total': queryset.count(),
            'pending': queryset.filter(status='PENDING').count(),
            'approved': queryset.filter(status='APPROVED').count(),
            'rejected': queryset.filter(status='REJECTED').count(),
            'by_type': {
                'accidents': queryset.filter(report_type='ACCIDENT').count(),
                'near_misses': queryset.filter(report_type='NEAR_MISS').count(),
                'potential_incidents': queryset.filter(report_type='POTENTIAL_INCIDENT').count(),
                'non_conformities': queryset.filter(report_type='NON_CONFORMITY').count(),
            },
            'by_severity': {
                'low': queryset.filter(severity='LOW').count(),
                'medium': queryset.filter(severity='MEDIUM').count(),
                'high': queryset.filter(severity='HIGH').count(),
                'critical': queryset.filter(severity='CRITICAL').count(),
            },
        }
        
        return Response(stats)


# =================================
# Form Template & Record Management
# =================================

class RecordViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing Records (submitted forms).
    
    Features:
    - Users can submit records
    - HSSE Managers/Admins can approve/reject
    - Auto-approval for Admin/HSSE Manager submissions
    - Year-based categorization and filtering
    - Email and in-app notifications
    """
    queryset = Record.objects.all().order_by('-created_at')
    serializer_class = RecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = Record.objects.all() if (user.position == 'HSSE MANAGER' or user.is_staff) else Record.objects.filter(submitted_by=user)
        
        # Filter by year
        year = self.request.query_params.get('year', None)
        if year:
            queryset = queryset.filter(year=year)
        
        # Filter by status
        status_filter = self.request.query_params.get('status', None)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        return queryset.order_by('-created_at')

    def perform_create(self, serializer):
        # The form_document_id is expected from the request data
        serializer.save(submitted_by=self.request.user)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def years(self, request):
        """Get list of years with record counts."""
        from django.db.models import Count, Q
        
        user = request.user
        queryset = Record.objects.all() if (user.position == 'HSSE MANAGER' or user.is_staff) else Record.objects.filter(submitted_by=user)
        
        years_data = queryset.values('year').annotate(
            total=Count('id'),
            pending=Count('id', filter=Q(status='PENDING_REVIEW')),
            approved=Count('id', filter=Q(status='APPROVED')),
            rejected=Count('id', filter=Q(status='REJECTED'))
        ).order_by('-year')
        
        return Response(list(years_data))

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def statistics(self, request):
        """Get record statistics."""
        user = request.user
        queryset = Record.objects.all() if (user.position == 'HSSE MANAGER' or user.is_staff) else Record.objects.filter(submitted_by=user)
        
        year = request.query_params.get('year', None)
        if year:
            queryset = queryset.filter(year=year)
        
        stats = {
            'total': queryset.count(),
            'pending': queryset.filter(status='PENDING_REVIEW').count(),
            'approved': queryset.filter(status='APPROVED').count(),
            'rejected': queryset.filter(status='REJECTED').count(),
            'by_year': list(queryset.values('year').annotate(count=Count('id')).order_by('-year')),
        }
        
        return Response(stats)

    @action(detail=True, methods=['post'], permission_classes=[IsHSSEManager])
    def approve(self, request, pk=None):
        """Approve a record and send notifications."""
        record = self.get_object()
        try:
            record.approve(request.user)
            return Response({
                'status': 'Record approved',
                'message': f'Record {record.record_number} approved successfully. Notification sent to submitter.'
            }, status=status.HTTP_200_OK)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], permission_classes=[IsHSSEManager])
    def reject(self, request, pk=None):
        """Reject a record with reason and send notifications."""
        record = self.get_object()
        serializer = RecordApprovalSerializer(data=request.data, context={'action': 'reject'})
        if serializer.is_valid():
            try:
                record.reject(request.user, reason=serializer.validated_data.get('rejection_reason', ''))
                return Response({
                    'status': 'Record rejected',
                    'message': f'Record {record.record_number} rejected. Notification sent to submitter.'
                }, status=status.HTTP_200_OK)
            except ValueError as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# LawCategory Views
class LawCategoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = LawCategory.objects.all()
    serializer_class = LawCategorySerializer
    permission_classes = [LegalCompliancePermission]

class LawCategoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LawCategory.objects.all()
    serializer_class = LawCategorySerializer
    permission_classes = [LegalCompliancePermission]

# LawResource Views
class LawResourceListCreateAPIView(generics.ListCreateAPIView):
    queryset = LawResource.objects.all()
    serializer_class = LawResourceSerializer
    permission_classes = [LegalCompliancePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['country', 'category', 'jurisdiction', 'is_repealed']
    search_fields = ['title', 'summary']

class LawResourceRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LawResource.objects.all()
    serializer_class = LawResourceSerializer
    permission_classes = [LegalCompliancePermission]

# LawResourceChange Views
class LawResourceChangeListCreateAPIView(generics.ListCreateAPIView):
    queryset = LawResourceChange.objects.all()
    serializer_class = LawResourceChangeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        if self.request.method in ['POST']:
            return [IsHSSEManager()]
        return super().get_permissions()

class LawResourceChangeRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LawResourceChange.objects.all()
    serializer_class = LawResourceChangeSerializer
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [IsHSSEManager()]
        return super().get_permissions()

# LegalRegisterEntry Views
class LegalRegisterEntryListCreateAPIView(generics.ListCreateAPIView):
    queryset = LegalRegisterEntry.objects.all()
    serializer_class = LegalRegisterEntrySerializer
    permission_classes = [LegalCompliancePermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['country', 'category', 'legislation_reference']
    search_fields = ['title', 'legal_obligation']

class LegalRegisterEntryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LegalRegisterEntry.objects.all()
    serializer_class = LegalRegisterEntrySerializer
    permission_classes = [LegalCompliancePermission]

# LegalRegisterComment Views
class LegalRegisterCommentListCreateAPIView(generics.ListCreateAPIView):
    queryset = LegalRegisterComment.objects.all()
    serializer_class = LegalRegisterCommentSerializer
    permission_classes = [LegalCompliancePermission]

class LegalRegisterCommentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LegalRegisterComment.objects.all()
    serializer_class = LegalRegisterCommentSerializer
    permission_classes = [LegalCompliancePermission]

# LegalRegisterDocument Views
class LegalRegisterDocumentListCreateAPIView(generics.ListCreateAPIView):
    queryset = LegalRegisterDocument.objects.all()
    serializer_class = LegalRegisterDocumentSerializer
    permission_classes = [LegalCompliancePermission]

class LegalRegisterDocumentRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LegalRegisterDocument.objects.all()
    serializer_class = LegalRegisterDocumentSerializer
    permission_classes = [LegalCompliancePermission]

# Position Views
class PositionListCreateAPIView(generics.ListCreateAPIView):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [LegalCompliancePermission]

class PositionRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer
    permission_classes = [LegalCompliancePermission]

# LegislationTracker Views
class LegislationTrackerListCreateAPIView(generics.ListCreateAPIView):
    queryset = LegislationTracker.objects.all()
    serializer_class = LegislationTrackerSerializer
    permission_classes = [LegalCompliancePermission]
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

class LegislationTrackerRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = LegislationTracker.objects.all()
    serializer_class = LegislationTrackerSerializer
    permission_classes = [LegalCompliancePermission]
    
    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

# =============================
# PPE Management Views
# =============================

class PPECategoryListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE categories."""
    queryset = PPECategory.objects.all()
    serializer_class = PPECategorySerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'description']

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class PPECategoryRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE categories."""
    queryset = PPECategory.objects.all()
    serializer_class = PPECategorySerializer
    permission_classes = [PPEManagementPermission]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class VendorListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating vendors."""
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_active', 'country']
    search_fields = ['name', 'contact_person', 'email']

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class VendorRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting vendors."""
    queryset = Vendor.objects.all()
    serializer_class = VendorSerializer
    permission_classes = [PPEManagementPermission]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class PPEPurchaseListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE purchases."""
    queryset = PPEPurchase.objects.all().order_by('-purchase_date')
    serializer_class = PPEPurchaseSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['vendor', 'ppe_category', 'status', 'received_by']
    search_fields = ['purchase_order_number', 'invoice_number', 'notes']
    
    def perform_create(self, serializer):
        serializer.save(received_by=self.request.user)

    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [IsHSSEManager()]
        return [IsAuthenticated()]


class PPEPurchaseRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE purchases."""
    queryset = PPEPurchase.objects.all()
    serializer_class = PPEPurchaseSerializer
    permission_classes = [PPEManagementPermission]


class PPEInventoryListAPIView(generics.ListAPIView):
    """API endpoint for listing PPE inventory."""
    queryset = PPEInventory.objects.all()
    serializer_class = PPEInventorySerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['ppe_category']
    search_fields = ['ppe_category__name']


class PPEInventoryRetrieveUpdateAPIView(generics.RetrieveUpdateAPIView):
    """API endpoint for retrieving and updating PPE inventory."""
    queryset = PPEInventory.objects.all()
    serializer_class = PPEInventorySerializer
    permission_classes = [IsHSSEManager]


class PPEIssueListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE issues."""
    queryset = PPEIssue.objects.all().order_by('-issue_date')
    serializer_class = PPEIssueSerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'ppe_category', 'issued_by']
    search_fields = ['employee__first_name', 'employee__last_name', 'ppe_category__name']

    def perform_create(self, serializer):
        serializer.save(issued_by=self.request.user)


class PPEIssueMyIssuesAPIView(generics.ListAPIView):
    """API endpoint for listing current user's PPE issues."""
    serializer_class = PPEIssueSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['ppe_category', 'status']
    search_fields = ['ppe_category__name']

    def get_queryset(self):
        """Filter to show only current user's PPE issues."""
        return PPEIssue.objects.filter(employee=self.request.user).order_by('-issue_date')


class PPEIssueRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE issues."""
    queryset = PPEIssue.objects.all()
    serializer_class = PPEIssueSerializer
    permission_classes = [PPEManagementPermission]


class PPERequestListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE requests."""
    queryset = PPERequest.objects.all().order_by('-created_at')
    serializer_class = PPERequestSerializer
    permission_classes = [PPEManagementPermission]  # Allow regular users to create requests
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'ppe_category', 'status']
    search_fields = ['employee__first_name', 'employee__last_name', 'ppe_category__name', 'reason']

    def perform_create(self, serializer):
        serializer.save(employee=self.request.user)

    def get_queryset(self):
        """Filter requests based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER':
            # Regular users can only see their own requests
            queryset = queryset.filter(employee=self.request.user)
        return queryset


class PPERequestRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE requests."""
    queryset = PPERequest.objects.all()
    serializer_class = PPERequestSerializer
    permission_classes = [PPEManagementPermission]  # Allow regular users to view own requests

    def get_queryset(self):
        """Filter requests based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see their own requests
            queryset = queryset.filter(employee=self.request.user)
        return queryset


class PPERequestApprovalAPIView(APIView):
    """API endpoint for approving/rejecting PPE requests."""
    permission_classes = [IsHSSEManager]

    def post(self, request, pk):
        try:
            ppe_request = get_object_or_404(PPERequest, pk=pk)
            serializer = PPERequestApprovalSerializer(data=request.data)
            
            if serializer.is_valid():
                status = serializer.validated_data['status']
                rejection_reason = serializer.validated_data.get('rejection_reason', '')
                
                ppe_request.status = status
                ppe_request.approved_by = request.user
                ppe_request.approved_at = timezone.now()
                
                if status == 'REJECTED':
                    ppe_request.rejection_reason = rejection_reason
                
                ppe_request.save()
                
                # If approved, create PPE issue
                if status == 'APPROVED':
                    PPEIssue.objects.create(
                        employee=ppe_request.employee,
                        ppe_category=ppe_request.ppe_category,
                        quantity=ppe_request.quantity,
                        issue_date=timezone.now().date(),
                        expiry_date=ppe_request.ppe_category.calculate_expiry_date(),
                        issued_by=request.user,
                        notes=f"Approved from request #{ppe_request.id}"
                    )
                
                return Response({'status': f'Request {status.lower()}'})
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except PPERequest.DoesNotExist:
            return Response({"error": "PPE request not found."}, status=404)


class PPEDamageReportListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE damage reports."""
    queryset = PPEDamageReport.objects.all().order_by('-reported_at')
    serializer_class = PPEDamageReportSerializer
    permission_classes = [PPEManagementPermission]  # Allow regular users to report damage
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'ppe_issue__ppe_category', 'is_approved', 'replacement_issued']
    search_fields = ['employee__first_name', 'employee__last_name', 'damage_description']

    def perform_create(self, serializer):
        serializer.save(employee=self.request.user)

    def get_queryset(self):
        """Filter reports based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see their own reports
            queryset = queryset.filter(employee=self.request.user)
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class PPEDamageReportRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE damage reports."""
    queryset = PPEDamageReport.objects.all()
    serializer_class = PPEDamageReportSerializer
    permission_classes = [PPEManagementPermission]  # Allow regular users to view own reports

    def get_queryset(self):
        """Filter reports based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see their own reports
            queryset = queryset.filter(employee=self.request.user)
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [IsHSSEManager()]
        return [IsAuthenticated()]


class PPEDamageReportReviewAPIView(APIView):
    """API endpoint for reviewing PPE damage reports."""
    permission_classes = [IsHSSEManager]

    def post(self, request, pk):
        try:
            damage_report = get_object_or_404(PPEDamageReport, pk=pk)
            serializer = PPEDamageReportReviewSerializer(data=request.data)
            
            if serializer.is_valid():
                is_approved = serializer.validated_data['is_approved']
                replacement_issued = serializer.validated_data.get('replacement_issued', False)
                
                damage_report.is_approved = is_approved
                damage_report.replacement_issued = replacement_issued
                damage_report.reviewed_by = request.user
                damage_report.reviewed_at = timezone.now()
                damage_report.save()
                
                return Response({'status': 'Damage report reviewed'})
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except PPEDamageReport.DoesNotExist:
            return Response({"error": "Damage report not found."}, status=404)


class PPETransferListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE transfers."""
    queryset = PPETransfer.objects.all().order_by('-transfer_date')
    serializer_class = PPETransferSerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['from_employee', 'to_employee', 'status']
    search_fields = ['from_employee__first_name', 'from_employee__last_name', 
                    'to_employee__first_name', 'to_employee__last_name']

    def perform_create(self, serializer):
        serializer.save(from_employee=self.request.user)

    def get_queryset(self):
        """Filter transfers based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see transfers involving them
            queryset = queryset.filter(
                Q(from_employee=self.request.user) | Q(to_employee=self.request.user)
            )
        return queryset


class PPETransferRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE transfers."""
    queryset = PPETransfer.objects.all()
    serializer_class = PPETransferSerializer
    permission_classes = [PPEManagementPermission]

    def get_queryset(self):
        """Filter transfers based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see transfers involving them
            queryset = queryset.filter(
                Q(from_employee=self.request.user) | Q(to_employee=self.request.user)
            )
        return queryset


class PPETransferApprovalAPIView(APIView):
    """API endpoint for approving/rejecting PPE transfers."""
    permission_classes = [IsHSSEManager]

    def post(self, request, pk):
        try:
            transfer = get_object_or_404(PPETransfer, pk=pk)
            serializer = PPETransferApprovalSerializer(data=request.data)
            
            if serializer.is_valid():
                status = serializer.validated_data['status']
                rejection_reason = serializer.validated_data.get('rejection_reason', '')
                
                transfer.status = status
                transfer.approved_by = request.user
                transfer.approved_at = timezone.now()
                
                if status == 'REJECTED':
                    transfer.rejection_reason = rejection_reason
                
                transfer.save()
                
                return Response({'status': f'Transfer {status.lower()}'})
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                
        except PPETransfer.DoesNotExist:
            return Response({"error": "PPE transfer not found."}, status=404)


class PPEReturnListCreateAPIView(generics.ListCreateAPIView):
    """API endpoint for listing and creating PPE returns."""
    queryset = PPEReturn.objects.all().order_by('-return_date')
    serializer_class = PPEReturnSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['employee', 'ppe_issue__ppe_category', 'received_by']
    search_fields = ['employee__first_name', 'employee__last_name', 'return_reason']

    def perform_create(self, serializer):
        serializer.save(employee=self.request.user, received_by=self.request.user)

    def get_queryset(self):
        """Filter returns based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER':
            # Regular users can only see their own returns
            queryset = queryset.filter(employee=self.request.user)
        return queryset


class PPEReturnRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    """API endpoint for retrieving, updating, and deleting PPE returns."""
    queryset = PPEReturn.objects.all()
    serializer_class = PPEReturnSerializer
    permission_classes = [PPEManagementPermission]

    def get_queryset(self):
        """Filter returns based on user role."""
        queryset = super().get_queryset()
        if self.request.user.position != 'HSSE MANAGER' and not self.request.user.is_superuser:
            # Regular users can only see their own returns
            queryset = queryset.filter(employee=self.request.user)
        return queryset


# =============================
# PPE Dashboard and Reporting Views
# =============================

class PPEStockPositionAPIView(APIView):
    """API endpoint for PPE stock position summary."""
    permission_classes = [PPEManagementPermission]

    def get(self, request):
        categories = PPECategory.objects.all()
        stock_data = []
        
        for category in categories:
            try:
                inventory = category.inventory
                total_received = PPEPurchase.objects.filter(
                    ppe_category=category
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                total_issued = PPEIssue.objects.filter(
                    ppe_category=category
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                total_damaged = PPEDamageReport.objects.filter(
                    ppe_issue__ppe_category=category,
                    is_approved=True
                ).count()
                
                # Filter expired issues in Python since is_expired is a property
                ppe_issues = PPEIssue.objects.filter(ppe_category=category)
                total_expired = sum(1 for issue in ppe_issues if issue.is_expired)
                
                stock_data.append({
                    'ppe_category': category,
                    'total_received': total_received,
                    'total_issued': total_issued,
                    'total_damaged': total_damaged,
                    'total_expired': total_expired,
                    'current_stock': inventory.current_stock,
                    'is_low_stock': inventory.is_low_stock
                })
            except PPEInventory.DoesNotExist:
                continue
        
        serializer = PPEStockPositionSerializer(stock_data, many=True, context={'request': request})
        return Response(serializer.data)


class PPEMovementAPIView(APIView):
    """API endpoint for PPE movement summary."""
    permission_classes = [PPEManagementPermission]

    def get(self, request):
        period = request.query_params.get('period', 'this_month')
        
        if period == 'this_month':
            start_date = timezone.now().replace(day=1)
        elif period == 'last_month':
            start_date = (timezone.now().replace(day=1) - timezone.timedelta(days=1)).replace(day=1)
        else:
            start_date = timezone.now().replace(day=1)
        
        end_date = timezone.now()
        
        categories = PPECategory.objects.all()
        movement_data = []
        
        for category in categories:
            total_received = PPEPurchase.objects.filter(
                ppe_category=category,
                purchase_date__range=[start_date, end_date]
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            total_issued = PPEIssue.objects.filter(
                ppe_category=category,
                issue_date__range=[start_date, end_date]
            ).aggregate(total=Sum('quantity'))['total'] or 0
            
            movement_data.append({
                'ppe_category': category,
                'total_received': total_received,
                'total_issued': total_issued,
                'movement_period': period.replace('_', ' ').title()
            })
        
        serializer = PPEMovementSerializer(movement_data, many=True, context={'request': request})
        return Response(serializer.data)


class PPEMostRequestedAPIView(APIView):
    """API endpoint for most requested PPE items."""
    permission_classes = [PPEManagementPermission]

    def get(self, request):
        period = request.query_params.get('period', '30')  # days
        
        start_date = timezone.now() - timezone.timedelta(days=int(period))
        
        # Get most requested PPE categories
        most_requested = PPERequest.objects.filter(
            created_at__gte=start_date
        ).values('ppe_category').annotate(
            request_count=Count('id'),
            total_quantity_requested=Sum('quantity')
        ).order_by('-request_count')[:10]
        
        data = []
        for item in most_requested:
            category = PPECategory.objects.get(id=item['ppe_category'])
            data.append({
                'ppe_category': category,
                'request_count': item['request_count'],
                'total_quantity_requested': item['total_quantity_requested']
            })
        
        serializer = PPEMostRequestedSerializer(data, many=True, context={'request': request})
        return Response(serializer.data)


class PPECostAnalysisAPIView(APIView):
    """API endpoint for PPE cost analysis."""
    permission_classes = [IsHSSEManager]

    def get(self, request):
        period = request.query_params.get('period', 'this_year')
        
        if period == 'this_year':
            start_date = timezone.now().replace(month=1, day=1)
        elif period == 'last_year':
            start_date = (timezone.now().replace(month=1, day=1) - timezone.timedelta(days=365))
        else:
            start_date = timezone.now().replace(day=1)
        
        end_date = timezone.now()
        
        categories = PPECategory.objects.all()
        cost_data = []
        
        for category in categories:
            purchases = PPEPurchase.objects.filter(
                ppe_category=category,
                purchase_date__range=[start_date, end_date]
            )
            
            total_cost = purchases.aggregate(total=Sum('total_cost'))['total'] or 0
            total_units = purchases.aggregate(total=Sum('quantity'))['total'] or 0
            avg_cost = total_cost / total_units if total_units > 0 else 0
            
            cost_data.append({
                'ppe_category': category,
                'total_cost': total_cost,
                'average_cost_per_unit': avg_cost,
                'total_units_purchased': total_units
            })
        
        serializer = PPECostAnalysisSerializer(cost_data, many=True, context={'request': request})
        return Response(serializer.data)


class PPEExpiryAlertsAPIView(APIView):
    """API endpoint for PPE expiry alerts."""
    permission_classes = [PPEManagementPermission]

    def get(self, request):
        days_threshold = int(request.query_params.get('days', 30))
        
        # Get PPE issues that are expiring soon or already expired
        expiring_issues = PPEIssue.objects.filter(
            expiry_date__lte=timezone.now().date() + timezone.timedelta(days=days_threshold)
        ).order_by('expiry_date')
        
        alerts = []
        for issue in expiring_issues:
            days_until_expiry = (issue.expiry_date - timezone.now().date()).days
            alerts.append({
                'ppe_issue': issue,
                'days_until_expiry': days_until_expiry,
                'is_expired': days_until_expiry < 0
            })
        
        serializer = PPEExpiryAlertSerializer(alerts, many=True, context={'request': request})
        return Response(serializer.data)


class PPELowStockAlertsAPIView(APIView):
    """API endpoint for PPE low stock alerts."""
    permission_classes = [PPEManagementPermission]

    def get(self, request):
        # Get all inventories and filter by low stock in Python
        all_inventories = PPEInventory.objects.all()
        low_stock_inventories = [inv for inv in all_inventories if inv.is_low_stock]
        
        alerts = []
        for inventory in low_stock_inventories:
            alerts.append({
                'ppe_category': inventory.ppe_category,
                'current_stock': inventory.current_stock,
                'threshold': inventory.ppe_category.low_stock_threshold
            })
        
        serializer = PPELowStockAlertSerializer(alerts, many=True, context={'request': request})
        return Response(serializer.data)


class PPEUserStockAPIView(APIView):
    """API endpoint for user's current PPE stock."""
    permission_classes = [PPEManagementPermission]

    def get(self, request, user_id=None):
        if user_id and request.user.position == 'HSSE MANAGER':
            user = get_object_or_404(User, id=user_id)
        else:
            user = request.user
        
        ppe_items = PPEIssue.objects.filter(
            employee=user,
            is_expired=False
        ).order_by('expiry_date')
        
        data = {
            'employee': user,
            'ppe_items': ppe_items,
            'total_items': ppe_items.count()
        }
        
        serializer = PPEUserStockSerializer(data, context={'request': request})
        return Response(serializer.data)


# =============================
# PPE Bulk Operations
# =============================

class BulkPPEIssueAPIView(APIView):
    """API endpoint for bulk PPE issuance."""
    permission_classes = [IsHSSEManager]

    def post(self, request):
        serializer = BulkPPEIssueSerializer(data=request.data)
        
        if serializer.is_valid():
            employee_ids = serializer.validated_data['employee_ids']
            ppe_category_id = serializer.validated_data['ppe_category_id']
            quantity_per_employee = serializer.validated_data['quantity_per_employee']
            issue_date = serializer.validated_data['issue_date']
            notes = serializer.validated_data.get('notes', '')
            
            try:
                ppe_category = PPECategory.objects.get(id=ppe_category_id)
                inventory = ppe_category.inventory
                
                # Check if there's enough stock
                total_quantity_needed = len(employee_ids) * quantity_per_employee
                if inventory.current_stock < total_quantity_needed:
                    return Response({
                        'error': f'Insufficient stock. Available: {inventory.current_stock}, Needed: {total_quantity_needed}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Create PPE issues for each employee
                created_issues = []
                for employee_id in employee_ids:
                    try:
                        employee = User.objects.get(id=employee_id)
                        issue = PPEIssue.objects.create(
                            employee=employee,
                            ppe_category=ppe_category,
                            quantity=quantity_per_employee,
                            issue_date=issue_date,
                            expiry_date=ppe_category.calculate_expiry_date(),
                            issued_by=request.user,
                            notes=notes
                        )
                        created_issues.append(issue)
                    except User.DoesNotExist:
                        continue
                
                return Response({
                    'message': f'Successfully issued PPE to {len(created_issues)} employees',
                    'created_count': len(created_issues)
                })
                
            except PPECategory.DoesNotExist:
                return Response({'error': 'PPE category not found'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BulkPPERequestApprovalAPIView(APIView):
    """API endpoint for bulk PPE request approval."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request):
        serializer = BulkPPERequestApprovalSerializer(data=request.data)
        if serializer.is_valid():
            request_ids = serializer.validated_data['request_ids']
            status = serializer.validated_data['status']
            rejection_reason = serializer.validated_data.get('rejection_reason', '')
            
            requests = PPERequest.objects.filter(id__in=request_ids)
            
            for req in requests:
                req.status = status
                if status == 'APPROVED':
                    req.approved_by = request.user
                    req.approved_at = timezone.now()
                elif status == 'REJECTED':
                    req.rejection_reason = rejection_reason
                req.save()
            
            return Response({
                'message': f'{len(requests)} requests {status.lower()} successfully'
            })
        return Response(serializer.errors, status=400)


class PPEPurchaseStatusUpdateAPIView(APIView):
    """API endpoint for updating purchase order status."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request, pk):
        try:
            purchase = PPEPurchase.objects.get(pk=pk)
            new_status = request.data.get('status')
            
            if new_status not in dict(PPEPurchase.STATUS_CHOICES):
                return Response({'error': 'Invalid status'}, status=400)
            
            purchase.status = new_status
            purchase.save()
            
            return Response({
                'message': f'Purchase order status updated to {new_status}',
                'status': new_status
            })
        except PPEPurchase.DoesNotExist:
            return Response({'error': 'Purchase order not found'}, status=404)


class PPEPurchaseReceiptCreateAPIView(APIView):
    """API endpoint for creating purchase receipts."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request, pk):
        try:
            purchase = PPEPurchase.objects.get(pk=pk)
            
            if not purchase.can_receive:
                return Response({
                    'error': 'Purchase order cannot be received. Status must be CONFIRMED or SHIPPED.'
                }, status=400)
            
            if hasattr(purchase, 'receipt'):
                return Response({
                    'error': 'Purchase order has already been received.'
                }, status=400)
            
            serializer = PPEPurchaseReceiptSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save(
                    purchase=purchase,
                    received_by=request.user
                )
                return Response({
                    'message': 'Purchase receipt created successfully',
                    'receipt': serializer.data
                })
            return Response(serializer.errors, status=400)
        except PPEPurchase.DoesNotExist:
            return Response({'error': 'Purchase order not found'}, status=404)


class PPEPurchaseReceiptListAPIView(generics.ListAPIView):
    """API endpoint for listing purchase receipts."""
    queryset = PPEPurchaseReceipt.objects.all().order_by('-received_date')
    serializer_class = PPEPurchaseReceiptSerializer
    permission_classes = [PPEManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['purchase__vendor', 'purchase__ppe_category', 'quality_check_passed']
    search_fields = ['purchase__purchase_order_number', 'notes']
    
    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'PATCH', 'DELETE']:
            return [IsHSSEManager()]
        return [IsAuthenticated()]

@api_view(['GET'])
@permission_classes([AllowAny])
def health_check(request):
    """
    Health check endpoint for monitoring
    """
    health_status = {
        'status': 'healthy',
        'timestamp': None,
        'database': 'unknown',
        'cache': 'unknown',
        'celery': 'unknown',
        'version': '1.0.0'
    }
    
    # Check database connection
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            health_status['database'] = 'connected'
    except Exception as e:
        health_status['database'] = f'error: {str(e)}'
        health_status['status'] = 'unhealthy'
    
    # Check cache (Redis)
    try:
        cache.set('health_check', 'ok', 10)
        if cache.get('health_check') == 'ok':
            health_status['cache'] = 'connected'
        else:
            health_status['cache'] = 'error'
            health_status['status'] = 'unhealthy'
    except Exception as e:
        health_status['cache'] = f'error: {str(e)}'
        health_status['status'] = 'unhealthy'
    
    # Check Celery (if available)
    try:
        from celery import current_app
        if current_app.control.inspect().active():
            health_status['celery'] = 'connected'
        else:
            health_status['celery'] = 'disconnected'
    except Exception as e:
        health_status['celery'] = f'error: {str(e)}'
    
    # Add timestamp
    from django.utils import timezone
    health_status['timestamp'] = timezone.now().isoformat()
    
    # Return appropriate status code
    if health_status['status'] == 'healthy':
        return Response(health_status, status=status.HTTP_200_OK)
    else:
        return Response(health_status, status=status.HTTP_503_SERVICE_UNAVAILABLE)

@api_view(['GET'])
@permission_classes([AllowAny])
def system_info(request):
    """
    System information endpoint (for debugging)
    """
    import psutil
    
    info = {
        'python_version': os.sys.version,
        'django_version': '5.0.14',
        'cpu_count': psutil.cpu_count(),
        'memory_total': psutil.virtual_memory().total,
        'memory_available': psutil.virtual_memory().available,
        'disk_usage': psutil.disk_usage('/').percent,
        'environment': 'production' if not settings.DEBUG else 'development',
    }
    
    return Response(info, status=status.HTTP_200_OK)


# =============================
# Audit Management Views
# =============================
from audits.models import (
    ISOClause45001, AuditPlan, AuditChecklist, AuditChecklistResponse,
    AuditFinding, CAPA, AuditEvidence, AuditReport, CAPAProgressUpdate,
    AuditMeeting, AuditComment
)
from audits.serializers import (
    ISOClause45001Serializer, AuditPlanListSerializer, AuditPlanDetailSerializer,
    AuditChecklistSerializer, AuditChecklistResponseSerializer,
    AuditFindingListSerializer, AuditFindingDetailSerializer,
    CAPAListSerializer, CAPADetailSerializer, AuditEvidenceSerializer,
    AuditReportSerializer, CAPAProgressUpdateSerializer, AuditMeetingSerializer,
    AuditCommentSerializer, AuditDashboardSerializer, BulkCAPAAssignSerializer
)
from datetime import datetime, timedelta
from decimal import Decimal


# ISO Clause Views
# Audit Type Views
from audits.models import AuditType, AuditChecklistTemplate, AuditScoringCriteria, AuditFinding, CompanySettings
from audits.serializers import AuditTypeSerializer, AuditChecklistTemplateSerializer, AuditScoringCriteriaSerializer, CompanySettingsSerializer

class AuditTypeListView(generics.ListAPIView):
    """List all active audit types."""
    serializer_class = AuditTypeSerializer
    permission_classes = [AuditManagementPermission]
    
    def get_queryset(self):
        """Get only active audit types."""
        return AuditType.objects.filter(is_active=True)


class AuditChecklistTemplateListView(generics.ListAPIView):
    """List all active checklist templates."""
    serializer_class = AuditChecklistTemplateSerializer
    permission_classes = [AuditManagementPermission]
    
    def get_queryset(self):
        """Get templates, optionally filter by audit type."""
        queryset = AuditChecklistTemplate.objects.filter(is_active=True)
        
        audit_type_id = self.request.query_params.get('audit_type_id', None)
        if audit_type_id:
            queryset = queryset.filter(audit_type_id=audit_type_id)
        
        return queryset


class AuditChecklistTemplateDetailView(generics.RetrieveAPIView):
    """Get detailed checklist template with all categories and questions."""
    serializer_class = AuditChecklistTemplateSerializer
    permission_classes = [AuditManagementPermission]
    queryset = AuditChecklistTemplate.objects.filter(is_active=True)


class AuditScoringCriteriaListView(generics.ListAPIView):
    """List all active scoring criteria."""
    serializer_class = AuditScoringCriteriaSerializer
    permission_classes = [AuditManagementPermission]
    queryset = AuditScoringCriteria.objects.filter(is_active=True)


class AuditScoreCalculationView(APIView):
    """Calculate audit score for a finding."""
    permission_classes = [AuditManagementPermission]
    
    def get(self, request, pk):
        """Calculate score for a finding."""
        try:
            finding = AuditFinding.objects.get(pk=pk)
        except AuditFinding.DoesNotExist:
            return Response(
                {'error': 'Finding not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        score_data = finding.calculate_overall_score()
        
        if score_data is None:
            return Response(
                {'error': 'No scoring template available for this audit type'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response(score_data, status=status.HTTP_200_OK)


class AuditFindingPDFReportView(APIView):
    """Generate PDF report for an audit finding."""
    permission_classes = [AuditManagementPermission]
    
    def get(self, request, pk):
        """Generate and return PDF report."""
        from django.http import HttpResponse
        from audits.pdf_report import generate_finding_pdf
        
        try:
            finding = AuditFinding.objects.select_related(
                'audit_plan', 'audit_plan__audit_type', 'audit_plan__lead_auditor',
                'iso_clause', 'identified_by'
            ).prefetch_related(
                'question_responses__question__category',
                'capas__responsible_person'
            ).get(pk=pk)
        except AuditFinding.DoesNotExist:
            return Response(
                {'error': 'Finding not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            # Generate PDF
            pdf_content = generate_finding_pdf(finding)
            
            # Create response
            response = HttpResponse(pdf_content, content_type='application/pdf')
            filename = f"Audit_Finding_{finding.finding_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Failed to generate PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ISOClause45001ListView(generics.ListCreateAPIView):
    """List and create ISO 45001 clauses."""
    queryset = ISOClause45001.objects.all()
    serializer_class = ISOClause45001Serializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_mandatory', 'risk_category']
    search_fields = ['clause_number', 'title', 'description']
    ordering_fields = ['clause_number']


class ISOClause45001DetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete ISO 45001 clause."""
    queryset = ISOClause45001.objects.all()
    serializer_class = ISOClause45001Serializer
    permission_classes = [IsAuthenticated]


# Audit Plan Views
class AuditPlanListCreateView(generics.ListCreateAPIView):
    """List and create audit plans. Only HSSE Manager can create."""
    queryset = AuditPlan.objects.all()
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['audit_type', 'status', 'lead_auditor']
    search_fields = ['audit_code', 'title', 'scope_description']
    ordering_fields = ['planned_start_date', 'created_at']
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return AuditPlanListSerializer
        return AuditPlanDetailSerializer
    
    def get_queryset(self):
        """Filter audits based on user role."""
        user = self.request.user
        if user.position == 'HSSE MANAGER' or user.is_staff:
            return AuditPlan.objects.all()
        # Other users see audits they're part of
        return AuditPlan.objects.filter(
            Q(audit_team=user) | Q(lead_auditor=user)
        ).distinct()
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create audits."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create audit plans.")
        serializer.save(created_by=self.request.user)


class AuditPlanDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete audit plan. Only HSSE Manager can modify."""
    queryset = AuditPlan.objects.all()
    serializer_class = AuditPlanDetailSerializer
    permission_classes = [AuditManagementPermission]
    
    def perform_update(self, serializer):
        """Only HSSE Manager can update."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can update audit plans.")
        serializer.save()
    
    def perform_destroy(self, instance):
        """Only HSSE Manager can delete."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can delete audit plans.")
        instance.delete()


# Audit Checklist Views
class AuditChecklistListCreateView(generics.ListCreateAPIView):
    """List and create checklist items."""
    serializer_class = AuditChecklistSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['audit_plan', 'iso_clause', 'question_type', 'is_mandatory']
    ordering_fields = ['sequence_order']
    
    def get_queryset(self):
        """Filter by audit plan if provided."""
        queryset = AuditChecklist.objects.all()
        audit_plan_id = self.request.query_params.get('audit_plan_id')
        if audit_plan_id:
            queryset = queryset.filter(audit_plan_id=audit_plan_id)
        return queryset
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create checklist items."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create checklist items.")
        serializer.save()


class AuditChecklistDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete checklist item."""
    queryset = AuditChecklist.objects.all()
    serializer_class = AuditChecklistSerializer
    permission_classes = [AuditManagementPermission]
    
    def perform_update(self, serializer):
        """Only HSSE Manager can update."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can update checklist items.")
        serializer.save()


# Audit Checklist Response Views
class AuditChecklistResponseListCreateView(generics.ListCreateAPIView):
    """List and create checklist responses."""
    serializer_class = AuditChecklistResponseSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['checklist_item', 'conformity_status', 'auditor']
    
    def get_queryset(self):
        """Filter by checklist item or audit plan."""
        queryset = AuditChecklistResponse.objects.all()
        checklist_item_id = self.request.query_params.get('checklist_item_id')
        audit_plan_id = self.request.query_params.get('audit_plan_id')
        
        if checklist_item_id:
            queryset = queryset.filter(checklist_item_id=checklist_item_id)
        if audit_plan_id:
            queryset = queryset.filter(checklist_item__audit_plan_id=audit_plan_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create responses."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can execute audits.")
        serializer.save(auditor=self.request.user)


class AuditChecklistResponseDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete checklist response."""
    queryset = AuditChecklistResponse.objects.all()
    serializer_class = AuditChecklistResponseSerializer
    permission_classes = [AuditManagementPermission]


# Audit Finding Views
class AuditFindingListCreateView(generics.ListCreateAPIView):
    """List and create audit findings."""
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['finding_type', 'severity', 'status', 'audit_plan', 'iso_clause']
    search_fields = ['finding_code', 'title', 'description', 'department_affected']
    ordering_fields = ['identified_date', 'severity', 'status']
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return AuditFindingListSerializer
        return AuditFindingDetailSerializer
    
    def get_queryset(self):
        """Filter findings based on user role."""
        user = self.request.user
        if user.position == 'HSSE MANAGER' or user.is_staff:
            return AuditFinding.objects.all()
        # Other users see findings from their department
        return AuditFinding.objects.filter(
            Q(department_affected=user.department) |
            Q(audit_plan__audit_team=user)
        ).distinct()
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create findings."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create findings.")
        serializer.save(identified_by=self.request.user)


class AuditFindingDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete finding."""
    queryset = AuditFinding.objects.all()
    serializer_class = AuditFindingDetailSerializer
    permission_classes = [AuditManagementPermission]
    
    def perform_update(self, serializer):
        """Only HSSE Manager can update."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can update findings.")
        serializer.save()


# CAPA Views
class CAPAListCreateView(generics.ListCreateAPIView):
    """List and create CAPAs."""
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['action_type', 'priority', 'status', 'responsible_person']
    search_fields = ['action_code', 'title', 'description']
    ordering_fields = ['target_completion_date', 'priority', 'created_at']
    
    def get_serializer_class(self):
        if self.request.method == 'GET':
            return CAPAListSerializer
        return CAPADetailSerializer
    
    def get_queryset(self):
        """Filter CAPAs based on user role."""
        user = self.request.user
        if user.position == 'HSSE MANAGER' or user.is_staff:
            return CAPA.objects.all()
        # Users see CAPAs assigned to them
        return CAPA.objects.filter(responsible_person=user)
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create CAPAs."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create CAPAs.")
        serializer.save(assigned_by=self.request.user)


class CAPADetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete CAPA."""
    queryset = CAPA.objects.all()
    serializer_class = CAPADetailSerializer
    permission_classes = [AuditManagementPermission]
    
    def perform_update(self, serializer):
        """Users can update their assigned CAPAs, HSSE Manager can update all."""
        capa = self.get_object()
        user = self.request.user
        
        if user.position == 'HSSE MANAGER' or user == capa.responsible_person:
            serializer.save()
        else:
            raise PermissionDenied("You can only update CAPAs assigned to you.")


# CAPA Progress Update View
class CAPAProgressUpdateCreateView(generics.CreateAPIView):
    """Create progress updates for CAPAs."""
    queryset = CAPAProgressUpdate.objects.all()
    serializer_class = CAPAProgressUpdateSerializer
    permission_classes = [AuditManagementPermission]
    
    def perform_create(self, serializer):
        """Users can update progress on their assigned CAPAs."""
        capa_id = self.request.data.get('capa')
        try:
            capa = CAPA.objects.get(id=capa_id)
            if self.request.user.position == 'HSSE MANAGER' or self.request.user == capa.responsible_person:
                # Update CAPA progress percentage
                progress_pct = self.request.data.get('progress_percentage')
                capa.progress_percentage = progress_pct
                capa.last_progress_update = timezone.now()
                capa.save()
                
                serializer.save(updated_by=self.request.user)
            else:
                raise PermissionDenied("You can only update CAPAs assigned to you.")
        except CAPA.DoesNotExist:
            raise ValidationError("CAPA not found.")


# Audit Evidence Views
class AuditEvidenceListCreateView(generics.ListCreateAPIView):
    """List and upload audit evidence."""
    queryset = AuditEvidence.objects.all()
    serializer_class = AuditEvidenceSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['file_type', 'audit_plan', 'finding', 'capa']
    
    def get_queryset(self):
        """Filter evidence based on access rights."""
        user = self.request.user
        queryset = AuditEvidence.objects.all()
        
        # Non-HSSE Manager can't see confidential evidence
        if user.position != 'HSSE MANAGER' and not user.is_staff:
            queryset = queryset.filter(is_confidential=False)
        
        # Filter by audit plan, finding, or CAPA if provided
        audit_plan_id = self.request.query_params.get('audit_plan_id')
        finding_id = self.request.query_params.get('finding_id')
        capa_id = self.request.query_params.get('capa_id')
        
        if audit_plan_id:
            queryset = queryset.filter(audit_plan_id=audit_plan_id)
        if finding_id:
            queryset = queryset.filter(finding_id=finding_id)
        if capa_id:
            queryset = queryset.filter(capa_id=capa_id)
        
        return queryset
    
    def perform_create(self, serializer):
        """Save evidence with uploader."""
        serializer.save(uploaded_by=self.request.user)


class AuditEvidenceDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete evidence."""
    queryset = AuditEvidence.objects.all()
    serializer_class = AuditEvidenceSerializer
    permission_classes = [AuditManagementPermission]


# Audit Report Views
class AuditReportListCreateView(generics.ListCreateAPIView):
    """List and create audit reports."""
    queryset = AuditReport.objects.all()
    serializer_class = AuditReportSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'audit_plan']
    
    def get_queryset(self):
        """HSSE Manager sees all, others see their department audits."""
        user = self.request.user
        if user.position == 'HSSE MANAGER' or user.is_staff:
            return AuditReport.objects.all()
        return AuditReport.objects.filter(
            Q(audit_plan__audit_team=user) |
            Q(audit_plan__lead_auditor=user)
        ).distinct()
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create reports."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create audit reports.")
        serializer.save(prepared_by=self.request.user)


class AuditReportDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete audit report."""
    queryset = AuditReport.objects.all()
    serializer_class = AuditReportSerializer
    permission_classes = [AuditManagementPermission]


# Audit Meeting Views
class AuditMeetingListCreateView(generics.ListCreateAPIView):
    """List and create audit meetings."""
    queryset = AuditMeeting.objects.all()
    serializer_class = AuditMeetingSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['audit_plan', 'meeting_type']
    
    def perform_create(self, serializer):
        """Only HSSE Manager can create meetings."""
        if self.request.user.position != 'HSSE MANAGER':
            raise PermissionDenied("Only HSSE Manager can create audit meetings.")
        serializer.save()


# Audit Comment Views
class AuditCommentListCreateView(generics.ListCreateAPIView):
    """List and create comments on audits/findings/CAPAs."""
    queryset = AuditComment.objects.all()
    serializer_class = AuditCommentSerializer
    permission_classes = [AuditManagementPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['audit_plan', 'finding', 'capa']
    
    def perform_create(self, serializer):
        """Save comment with author."""
        serializer.save(author=self.request.user)


# Audit Dashboard View
class AuditDashboardView(APIView):
    """Comprehensive audit dashboard with metrics and analytics."""
    permission_classes = [AuditManagementPermission]
    
    def get(self, request):
        """Get dashboard data."""
        user = request.user
        current_year = datetime.now().year
        
        # Base querysets
        if user.position == 'HSSE MANAGER' or user.is_staff:
            audits = AuditPlan.objects.all()
            findings = AuditFinding.objects.all()
            capas = CAPA.objects.all()
        else:
            # Limited view for non-HSSE users
            audits = AuditPlan.objects.filter(
                Q(audit_team=user) | Q(lead_auditor=user)
            ).distinct()
            findings = AuditFinding.objects.filter(
                Q(audit_plan__audit_team=user) | Q(department_affected=user.department)
            ).distinct()
            capas = CAPA.objects.filter(responsible_person=user)
        
        # Overall metrics
        total_audits = audits.count()
        audits_this_year = audits.filter(created_at__year=current_year).count()
        completed_audits = audits.filter(status='COMPLETED').count()
        in_progress_audits = audits.filter(status='IN_PROGRESS').count()
        scheduled_audits = audits.filter(status='SCHEDULED').count()
        
        # Findings metrics
        total_findings = findings.count()
        open_findings = findings.exclude(status='CLOSED').count()
        major_ncs = findings.filter(finding_type='MAJOR_NC').count()
        minor_ncs = findings.filter(finding_type='MINOR_NC').count()
        observations = findings.filter(finding_type='OBSERVATION').count()
        
        # CAPA metrics
        total_capas = capas.count()
        open_capas = capas.exclude(status='CLOSED').count()
        overdue_capas = sum(1 for capa in capas if capa.is_overdue)
        closed_capas = capas.filter(status='CLOSED').count()
        capa_completion_rate = (closed_capas / total_capas * 100) if total_capas > 0 else 0
        
        # Compliance scoring
        reports = AuditReport.objects.filter(audit_plan__in=audits, status='APPROVED')
        if reports.exists():
            avg_score = reports.aggregate(avg=Avg('overall_conformity_score'))['avg'] or 0
        else:
            avg_score = 0
        
        # Compliance trend (last 6 months)
        six_months_ago = timezone.now() - timedelta(days=180)
        trend_reports = reports.filter(report_date__gte=six_months_ago).order_by('report_date')
        compliance_trend = [
            {
                'month': report.report_date.strftime('%b %Y'),
                'score': float(report.overall_conformity_score)
            }
            for report in trend_reports
        ]
        
        # Findings by ISO clause
        findings_by_clause = {}
        for finding in findings:
            clause = finding.iso_clause.clause_number
            findings_by_clause[clause] = findings_by_clause.get(clause, 0) + 1
        
        # Compliance by clause (from latest reports)
        compliance_by_clause = {}
        latest_report = reports.order_by('-report_date').first()
        if latest_report and latest_report.conformity_by_clause:
            compliance_by_clause = latest_report.conformity_by_clause
        
        # Upcoming audits
        upcoming_audits = audits.filter(
            status__in=['SCHEDULED', 'IN_PROGRESS'],
            planned_start_date__gte=datetime.now().date()
        ).order_by('planned_start_date')[:5]
        
        # Recent findings
        recent_findings = findings.order_by('-identified_date')[:10]
        
        # Overdue CAPAs list
        overdue_capas_list = [capa for capa in capas if capa.is_overdue][:10]
        
        dashboard_data = {
            'total_audits': total_audits,
            'audits_this_year': audits_this_year,
            'completed_audits': completed_audits,
            'in_progress_audits': in_progress_audits,
            'scheduled_audits': scheduled_audits,
            
            'total_findings': total_findings,
            'open_findings': open_findings,
            'major_ncs': major_ncs,
            'minor_ncs': minor_ncs,
            'observations': observations,
            
            'total_capas': total_capas,
            'open_capas': open_capas,
            'overdue_capas': overdue_capas,
            'capa_completion_rate': round(capa_completion_rate, 2),
            
            'average_compliance_score': round(float(avg_score), 2),
            'compliance_trend': compliance_trend,
            
            'findings_by_clause': findings_by_clause,
            'compliance_by_clause': compliance_by_clause,
            
            'upcoming_audits': AuditPlanListSerializer(upcoming_audits, many=True).data,
            'recent_findings': AuditFindingListSerializer(recent_findings, many=True).data,
            'overdue_capas_list': CAPAListSerializer(overdue_capas_list, many=True).data,
        }
        
        return Response(dashboard_data, status=status.HTTP_200_OK)


# CAPA Bulk Operations
class BulkCAPAAssignView(APIView):
    """Bulk assign CAPAs to findings."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request):
        """Create CAPAs for multiple findings at once."""
        serializer = BulkCAPAAssignSerializer(data=request.data)
        
        if serializer.is_valid():
            finding_ids = serializer.validated_data['finding_ids']
            responsible_person = serializer.validated_data['responsible_person_id']
            target_date = serializer.validated_data['target_date']
            action_plan = serializer.validated_data['action_plan']
            
            created_capas = []
            for finding_id in finding_ids:
                try:
                    finding = AuditFinding.objects.get(id=finding_id)
                    capa = CAPA.objects.create(
                        finding=finding,
                        action_type='CORRECTIVE',
                        title=f"CAPA for {finding.title}",
                        description=action_plan,
                        root_cause=finding.root_cause_analysis.get('summary', ''),
                        action_plan=action_plan,
                        responsible_person=responsible_person,
                        assigned_by=request.user,
                        target_completion_date=target_date,
                        priority='HIGH' if finding.severity in ['CRITICAL', 'HIGH'] else 'MEDIUM',
                        effectiveness_criteria="To be defined"
                    )
                    created_capas.append(capa)
                    
                    # Update finding status
                    finding.status = 'CAPA_ASSIGNED'
                    finding.save()
                    
                except AuditFinding.DoesNotExist:
                    continue
            
            return Response({
                'message': f'{len(created_capas)} CAPAs created successfully',
                'capas': CAPAListSerializer(created_capas, many=True).data
            }, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# My CAPAs View
class MyCAPAsView(generics.ListAPIView):
    """Get CAPAs assigned to current user."""
    serializer_class = CAPAListSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['target_completion_date', 'priority']
    
    def get_queryset(self):
        """Get CAPAs for current user."""
        return CAPA.objects.filter(responsible_person=self.request.user)


# Email Notification Views
from audits.services import (
    send_audit_plan_notification, 
    send_capa_assignment_notification,
    send_finding_notification,
    send_audit_completion_notification
)


class SendAuditPlanNotificationView(APIView):
    """Send audit plan notification email to specified recipients."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request, pk):
        """
        Send audit plan notification.
        
        Request body:
        {
            "recipient_emails": ["email1@example.com", "email2@example.com"],  // optional
            "recipient_user_ids": [1, 2, 3],  // optional
            "additional_message": "Please prepare all HSE documents"  // optional
        }
        """
        try:
            audit_plan = AuditPlan.objects.get(pk=pk)
        except AuditPlan.DoesNotExist:
            return Response(
                {'error': 'Audit plan not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get recipients from request
        recipients = []
        
        # Email addresses
        recipient_emails = request.data.get('recipient_emails', [])
        recipients.extend(recipient_emails)
        
        # User IDs
        recipient_user_ids = request.data.get('recipient_user_ids', [])
        if recipient_user_ids:
            users = User.objects.filter(id__in=recipient_user_ids)
            recipients.extend(users)
        
        # If no recipients specified, use default (audit team)
        if not recipients:
            recipients = None
        
        additional_message = request.data.get('additional_message', '')
        
        # Send notification
        count = send_audit_plan_notification(audit_plan, recipients, additional_message)
        
        if count > 0:
            return Response({
                'message': f'Audit plan notification sent to {count} recipient(s)',
                'recipients_count': count
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'Failed to send notifications. No valid recipients found.'
            }, status=status.HTTP_400_BAD_REQUEST)


class SendCAPANotificationView(APIView):
    """Send CAPA assignment notification."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request, pk):
        """Send CAPA assignment notification to responsible person."""
        try:
            capa = CAPA.objects.get(pk=pk)
        except CAPA.DoesNotExist:
            return Response(
                {'error': 'CAPA not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        success = send_capa_assignment_notification(capa)
        
        if success:
            return Response({
                'message': f'CAPA notification sent to {capa.responsible_person.get_full_name}'
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'Failed to send CAPA notification'
            }, status=status.HTTP_400_BAD_REQUEST)


class SendFindingNotificationView(APIView):
    """Send finding notification to department and HSSE Manager."""
    permission_classes = [IsHSSEManager]
    
    def post(self, request, pk):
        """Send finding notification."""
        try:
            finding = AuditFinding.objects.get(pk=pk)
        except AuditFinding.DoesNotExist:
            return Response(
                {'error': 'Finding not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        success = send_finding_notification(finding)
        
        if success:
            return Response({
                'message': 'Finding notification sent successfully'
            }, status=status.HTTP_200_OK)
        else:
            return Response({
                'error': 'Failed to send finding notification'
            }, status=status.HTTP_400_BAD_REQUEST)


# ===========================================
# Company Settings Views
# ===========================================
class CompanySettingsView(APIView):
    """Get or update company settings (logo, company info)."""
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get(self, request):
        """Get company settings."""
        settings = CompanySettings.get_settings()
        serializer = CompanySettingsSerializer(settings, context={'request': request})
        return Response(serializer.data)
    
    def put(self, request):
        """Update company settings (HSSE Manager only)."""
        if request.user.position != 'HSSE MANAGER':
            return Response(
                {'error': 'Only HSSE Managers can update company settings'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        settings = CompanySettings.get_settings()
        serializer = CompanySettingsSerializer(
            settings, 
            data=request.data, 
            partial=True,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save(updated_by=request.user)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ===========================================
# RISK MANAGEMENT VIEWS
# ===========================================
from risks.models import (
    RiskAssessment, RiskHazard, RiskExposure, ControlBarrier,
    RiskTreatmentAction, RiskReview, RiskAttachment, RiskMatrixConfig
)
from risks.serializers import (
    RiskAssessmentListSerializer, RiskAssessmentDetailSerializer,
    RiskHazardSerializer, RiskExposureSerializer, ControlBarrierSerializer,
    RiskTreatmentActionSerializer, RiskReviewSerializer,
    RiskAttachmentSerializer, RiskMatrixConfigSerializer
)
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class RiskMatrixConfigView(APIView):
    """Get risk matrix configuration."""
    permission_classes = [RiskManagementPermission]  # Allow read-only for regular users
    
    def get(self, request):
        """Get risk matrix configuration."""
        config = RiskMatrixConfig.get_config()
        serializer = RiskMatrixConfigSerializer(config)
        return Response(serializer.data)


class RiskAssessmentListCreateView(generics.ListCreateAPIView):
    """List all risk assessments or create a new one."""
    permission_classes = [RiskManagementPermission]  # Allow read-only for regular users
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return RiskAssessmentDetailSerializer
        return RiskAssessmentListSerializer
    
    def get_queryset(self):
        queryset = RiskAssessment.objects.all().select_related(
            'assessed_by', 'approved_by', 'risk_owner'
        ).prefetch_related('hazards', 'barriers', 'treatment_actions')
        
        # Filter by status
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by risk category
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(risk_category=category)
        
        # Filter by risk level
        risk_level = self.request.query_params.get('risk_level')
        if risk_level:
            if risk_level == 'LOW':
                queryset = [ra for ra in queryset if ra.residual_risk_level <= 5]
            elif risk_level == 'MEDIUM':
                queryset = [ra for ra in queryset if 6 <= ra.residual_risk_level <= 12]
            elif risk_level == 'HIGH':
                queryset = [ra for ra in queryset if ra.residual_risk_level >= 15]
        
        # Filter by location
        location = self.request.query_params.get('location')
        if location:
            queryset = queryset.filter(location__icontains=location)
        
        return queryset


class RiskAssessmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Retrieve, update, or delete a risk assessment."""
    queryset = RiskAssessment.objects.all().select_related(
        'assessed_by', 'approved_by', 'risk_owner'
    ).prefetch_related(
        'hazards', 'barriers', 'treatment_actions', 'reviews', 'attachments'
    )
    serializer_class = RiskAssessmentDetailSerializer
    permission_classes = [RiskManagementPermission]  # Allow read-only for regular users
    
    def perform_update(self, serializer):
        risk_assessment = serializer.save()
        
        # If approved, set dates
        if risk_assessment.status == 'APPROVED' and not risk_assessment.approval_date:
            risk_assessment.approval_date = date.today()
            risk_assessment.approved_by = self.request.user
            risk_assessment.save()


class RiskAssessmentApproveView(APIView):
    """Approve a risk assessment (HSSE Manager only)."""
    permission_classes = [IsHSSEManager]  # Only HSSE Manager can approve
    
    def post(self, request, pk):
        risk_assessment = get_object_or_404(RiskAssessment, pk=pk)
        
        risk_assessment.status = 'APPROVED'
        risk_assessment.approved_by = request.user
        risk_assessment.approval_date = date.today()
        risk_assessment.save()
        
        serializer = RiskAssessmentDetailSerializer(risk_assessment, context={'request': request})
        return Response(serializer.data)


class RiskDashboardView(APIView):
    """Risk dashboard with analytics."""
    permission_classes = [RiskManagementPermission]  # Dashboard access control
    
    def get(self, request):
        from django.db.models import Count, Q
        
        # Total counts
        total_assessments = RiskAssessment.objects.count()
        active_assessments = RiskAssessment.objects.filter(status='ACTIVE').count()
        
        # Risk level distribution
        risk_assessments = RiskAssessment.objects.filter(status__in=['APPROVED', 'ACTIVE'])
        
        low_risk = sum(1 for ra in risk_assessments if ra.residual_risk_level <= 5)
        medium_risk = sum(1 for ra in risk_assessments if 6 <= ra.residual_risk_level <= 12)
        high_risk = sum(1 for ra in risk_assessments if ra.residual_risk_level >= 15)
        
        # Overdue reviews
        overdue_reviews = sum(1 for ra in risk_assessments if ra.is_overdue_for_review)
        
        # Risk by category
        by_category = {}
        for category, _ in RiskAssessment.RISK_CATEGORY_CHOICES:
            count = RiskAssessment.objects.filter(
                risk_category=category,
                status__in=['APPROVED', 'ACTIVE']
            ).count()
            by_category[category] = count
        
        # Pending actions
        pending_actions = RiskTreatmentAction.objects.filter(
            status__in=['PLANNED', 'IN_PROGRESS']
        ).count()
        
        overdue_actions = sum(
            1 for action in RiskTreatmentAction.objects.filter(status__in=['PLANNED', 'IN_PROGRESS'])
            if action.is_overdue
        )
        
        return Response({
            'total_assessments': total_assessments,
            'active_assessments': active_assessments,
            'risk_distribution': {
                'low': low_risk,
                'medium': medium_risk,
                'high': high_risk,
            },
            'overdue_reviews': overdue_reviews,
            'by_category': by_category,
            'pending_actions': pending_actions,
            'overdue_actions': overdue_actions,
        })


class RiskExcelExportView(APIView):
    """Export risk assessments to Excel."""
    permission_classes = [RiskManagementPermission]  # Export access control
    
    def get(self, request):
        """Export all or filtered risk assessments to Excel."""
        from datetime import datetime
        
        # Get filter parameters
        status_filter = request.query_params.get('status')
        category_filter = request.query_params.get('category')
        
        queryset = RiskAssessment.objects.all().select_related(
            'assessed_by', 'risk_owner'
        ).prefetch_related('hazards', 'barriers', 'treatment_actions')
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if category_filter:
            queryset = queryset.filter(risk_category=category_filter)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risk Register"
        
        # Header styling
        header_fill = PatternFill(start_color="0052D4", end_color="0052D4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Define headers
        headers = [
            'Event Number', 'Status', 'Location', 'Process Area', 'Activity Type',
            'Risk Category', 'Hazard Type', 'Hazard Description', 'Event Description',
            'Causes', 'Consequences', 'Initial Prob', 'Initial Sev', 'Initial Risk',
            'Initial Rating', 'Preventive Barriers', 'Protective Barriers',
            'Residual Prob', 'Residual Sev', 'Residual Risk', 'Residual Rating',
            'Risk Acceptable', 'ALARP Required', 'Additional Actions',
            'Person Responsible', 'Target Date', 'Assessed By', 'Assessment Date',
            'Next Review', 'Comments'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Write data
        row = 2
        for ra in queryset:
            hazards = ra.hazards.all()
            hazard = hazards.first() if hazards.exists() else None
            
            preventive_barriers = "; ".join([
                b.description for b in ra.barriers.filter(barrier_type='PREVENTIVE')
            ])
            protective_barriers = "; ".join([
                b.description for b in ra.barriers.filter(barrier_type='PROTECTIVE')
            ])
            additional_actions = "; ".join([
                a.action_description for a in ra.treatment_actions.all()
            ])
            
            data = [
                ra.event_number,
                ra.status,
                ra.location,
                ra.process_area,
                ra.get_activity_type_display(),
                ra.get_risk_category_display(),
                hazard.get_hazard_type_display() if hazard else '',
                hazard.hazard_description if hazard else '',
                hazard.event_description if hazard else '',
                hazard.causes if hazard else '',
                hazard.consequences if hazard else '',
                ra.initial_probability,
                ra.initial_severity,
                ra.initial_risk_level,
                ra.initial_risk_rating,
                preventive_barriers,
                protective_barriers,
                ra.residual_probability,
                ra.residual_severity,
                ra.residual_risk_level,
                ra.residual_risk_rating,
                'Yes' if ra.risk_acceptable else 'No',
                'Yes' if ra.alarp_required else 'No',
                additional_actions,
                ra.risk_owner.get_full_name if ra.risk_owner else '',
                ra.next_review_date,
                ra.assessed_by.get_full_name if ra.assessed_by else '',
                ra.assessment_date,
                ra.next_review_date,
                ra.comments,
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # Color code risk levels
                if col == 14:  # Initial risk level
                    cell.fill = PatternFill(
                        start_color=ra.initial_risk_color.replace('#', ''),
                        end_color=ra.initial_risk_color.replace('#', ''),
                        fill_type="solid"
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
                elif col == 20:  # Residual risk level
                    cell.fill = PatternFill(
                        start_color=ra.residual_risk_color.replace('#', ''),
                        end_color=ra.residual_risk_color.replace('#', ''),
                        fill_type="solid"
                    )
                    cell.font = Font(color="FFFFFF", bold=True)
            
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'Risk_Register_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


class MyRiskAssessmentsView(generics.ListAPIView):
    """Get risk assessments created by or assigned to current user."""
    serializer_class = RiskAssessmentListSerializer
    permission_classes = [RiskManagementPermission]
    
    def get_queryset(self):
        user = self.request.user
        return RiskAssessment.objects.filter(
            Q(assessed_by=user) | Q(risk_owner=user)
        ).select_related('assessed_by', 'risk_owner').prefetch_related('hazards', 'barriers')
